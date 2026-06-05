import io
import os
import tempfile
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.equipment import Equipment
from app.models.equipment_group import EquipmentGroup
from app.models.maintenance_record import MaintenanceRecord
from app.models.plant import Plant
from app.models.user import User
from app.security import get_current_user
from app.services.excel_importer import get_sheet_names, parse_excel_maintenance_records
from app.services.audit_service import log_action

router = APIRouter()


@router.get("/template")
def download_template(current_user: User = Depends(get_current_user)):
    """Return a pre-formatted Excel template for bulk maintenance record import."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Maintenance Records"

    # ── Column definitions ────────────────────────────────────────────────────
    columns = [
        ("Date",              "Required. Format: YYYY-MM-DD or DD/MM/YYYY.  e.g. 2026-06-15",          15),
        ("Time Reported",     "Optional. Format: HH:MM (24-hour).  e.g. 08:30",                        14),
        ("MR No",             "Optional. Unique maintenance request number.  e.g. MR-0042",            14),
        ("Plant",             "Optional. Must match an existing plant name or will be created.",        20),
        ("Equipment Group",   "Optional. Group within the plant.  e.g. Electrical",                    20),
        ("Equipment Name",    "Optional. Asset / machine name.  e.g. Main Conveyor Drive",             28),
        ("Fault Description", "Required. Describe the issue or fault.",                                 36),
        ("Artisan Name",      "Optional. Technician who carried out the work.",                         20),
        ("Reported By",       "Optional. Person who logged the fault.",                                 20),
        ("Reported To",       "Optional. Supervisor or foreman notified.",                              20),
        ("Arrival Time",      "Optional. Format: HH:MM (24-hour).  e.g. 09:15",                        14),
        ("Finishing Time",    "Optional. Format: HH:MM (24-hour).  e.g. 11:45",                        14),
        ("Downtime",          "Optional. Duration in minutes, hours (e.g. 90) or HH:MM (e.g. 1:30).",  14),
        ("Remarks",           "Optional. Additional notes or observations.",                             30),
        ("Record Type",       "Optional. Enter  regular  or  breakdown  (default: regular).",           16),
    ]

    header_fill   = PatternFill("solid", fgColor="1E293B")
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    note_fill     = PatternFill("solid", fgColor="F1F5F9")
    note_font     = Font(italic=True, color="475569", size=9)
    example_font  = Font(color="374151", size=10)
    required_fill = PatternFill("solid", fgColor="FEF2F2")

    # ── Row 1: column headers ─────────────────────────────────────────────────
    for col_idx, (name, _, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Row 2: field notes ────────────────────────────────────────────────────
    for col_idx, (_, note, _) in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=note)
        cell.font  = note_font
        cell.fill  = note_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 32

    # ── Rows 3-5: example data ────────────────────────────────────────────────
    examples = [
        ["2026-06-15", "07:45", "MR-0001", "Main Bakery", "Mechanical", "Conveyor Drive Unit",
         "Conveyor belt slipping on startup, causing product spillage.", "J. Dlamini",
         "S. Mokoena", "Plant Manager", "08:00", "10:30", "150", "Replaced worn tensioner pulley.", "breakdown"],
        ["2026-06-16", "09:00", "MR-0002", "Packaging Line", "Electrical", "Labelling Machine",
         "Label sensor misaligned — triggering false rejects.", "T. Nkosi",
         "A. Sithole", "Shift Supervisor", "09:15", "10:00", "45", "Realigned sensor bracket.", "regular"],
        ["2026-06-17", "",      "MR-0003", "Dispatch",       "",            "Pallet Wrapper",
         "Motor not starting. Thermal overload tripped.", "J. Dlamini",
         "B. Khumalo", "", "14:00", "15:30", "90", "", "breakdown"],
    ]

    for row_idx, row_data in enumerate(examples, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = example_font
            cell.alignment = Alignment(vertical="center", wrap_text=False)
        ws.row_dimensions[row_idx].height = 18

    # Highlight required columns (Date + Fault Description) in example rows
    for row_idx in range(3, 6):
        for col_idx in (1, 7):
            ws.cell(row=row_idx, column=col_idx).fill = required_fill

    # Freeze the header rows so they stay visible when scrolling
    ws.freeze_panes = "A3"

    # ── Instructions sheet ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ("MAINTENANCE RECORDS IMPORT TEMPLATE", True, 14),
        ("", False, 11),
        ("REQUIRED COLUMNS", True, 11),
        ("  • Date            — Record date (YYYY-MM-DD or DD/MM/YYYY)", False, 10),
        ("  • Fault Description — Describe the fault or issue", False, 10),
        ("", False, 10),
        ("OPTIONAL COLUMNS", True, 11),
        ("  • Time Reported   — HH:MM 24-hour format", False, 10),
        ("  • MR No           — Unique reference number", False, 10),
        ("  • Plant           — Must match an existing plant name (created if missing)", False, 10),
        ("  • Equipment Group — Must match a group in that plant", False, 10),
        ("  • Equipment Name  — Must match an existing equipment name", False, 10),
        ("  • Artisan Name    — Technician who did the work", False, 10),
        ("  • Reported By     — Person who logged the fault", False, 10),
        ("  • Reported To     — Supervisor notified", False, 10),
        ("  • Arrival Time    — HH:MM 24-hour format", False, 10),
        ("  • Finishing Time  — HH:MM 24-hour format (downtime auto-calculated)", False, 10),
        ("  • Downtime        — Minutes, hours, or HH:MM (e.g. 90, 1.5, or 1:30)", False, 10),
        ("  • Remarks         — Additional notes", False, 10),
        ("  • Record Type     — 'regular' or 'breakdown' (default: regular)", False, 10),
        ("", False, 10),
        ("NOTES", True, 11),
        ("  • Row 1 is the header — do not delete or rename the headers.", False, 10),
        ("  • Row 2 contains field descriptions — you may delete it before importing.", False, 10),
        ("  • Delete the example rows (3-5) before importing real data.", False, 10),
        ("  • If a record with the same MR No already exists it will be updated.", False, 10),
        ("  • Status is set to 'closed' for all imported records.", False, 10),
        ("  • Column order does not matter — the importer matches by name.", False, 10),
    ]
    ws2.column_dimensions["A"].width = 72
    for row_idx, (text, bold, size) in enumerate(instructions, start=1):
        cell = ws2.cell(row=row_idx, column=1, value=text)
        cell.font = Font(bold=bold, size=size, color="1E293B" if bold else "374151")
        cell.alignment = Alignment(vertical="center")
        ws2.row_dimensions[row_idx].height = 18 if text else 8

    # ── Stream the workbook ───────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="maintenance_import_template.xlsx"'},
    )


@router.post("/preview")
async def preview_import(
    file: UploadFile = File(...),
    sheet_name: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls files are accepted")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        sheets = get_sheet_names(tmp_path)
        selected_sheet = sheet_name or sheets[0]
        records = parse_excel_maintenance_records(tmp_path, sheet_name=selected_sheet)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    finally:
        os.unlink(tmp_path)

    return {
        "sheets": sheets,
        "selected_sheet": selected_sheet,
        "total_records": len(records),
        "preview": records[:20],
    }


@router.post("/commit")
async def commit_import(
    file: UploadFile = File(...),
    sheet_name: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls files are accepted")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        sheets = get_sheet_names(tmp_path)
        selected_sheet = sheet_name or sheets[0]
        records = parse_excel_maintenance_records(tmp_path, sheet_name=selected_sheet)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    finally:
        os.unlink(tmp_path)

    saved = 0
    created = 0
    updated = 0
    errors = []

    for i, row in enumerate(records):
        try:
            plant_name = row.pop("plant", None)
            equipment_name = row.pop("equipment", None)
            equipment_group_name = row.pop("equipment_group", None)

            plant_id = None
            if plant_name:
                plant = db.query(Plant).filter(Plant.name.ilike(plant_name)).first()
                if not plant:
                    plant = Plant(name=plant_name.strip().title())
                    db.add(plant)
                    db.flush()
                plant_id = plant.id

            equipment_id = None
            equip = None
            if equipment_name:
                equip = db.query(Equipment).filter(
                    Equipment.equipment_name.ilike(equipment_name)
                ).first()
                if not equip:
                    equip = Equipment(
                        equipment_name=equipment_name.strip(),
                        plant_id=plant_id,
                    )
                    db.add(equip)
                    db.flush()
                equipment_id = equip.id

            group_id = None
            if equipment_group_name:
                group_query = db.query(EquipmentGroup).filter(EquipmentGroup.name.ilike(equipment_group_name.strip()))
                if plant_id is not None:
                    group_query = group_query.filter(EquipmentGroup.plant_id == plant_id)
                group = group_query.first()
                if not group:
                    group = EquipmentGroup(name=equipment_group_name.strip(), plant_id=plant_id)
                    db.add(group)
                    db.flush()
                group_id = group.id

            if equip is not None and group_id is None and equip.equipment_group_id is not None:
                group_id = equip.equipment_group_id

            if equip is not None and group_id is not None:
                equip.equipment_group_id = group_id

            existing_record = None
            incoming_mr_no = row.get("mr_no")

            if incoming_mr_no:
                existing_record = (
                    db.query(MaintenanceRecord)
                    .filter(MaintenanceRecord.mr_no.ilike(str(incoming_mr_no).strip()))
                    .first()
                )

            if not existing_record:
                existing_record = (
                    db.query(MaintenanceRecord)
                    .filter(
                        MaintenanceRecord.record_date == row.get("record_date"),
                        MaintenanceRecord.plant_id == plant_id,
                        MaintenanceRecord.equipment_id == equipment_id,
                        MaintenanceRecord.issue_description == row.get("issue_description"),
                    )
                    .first()
                )

            # Ensure group fields are not duplicated in the incoming data
            row.pop("equipment_group", None)
            row.pop("equipment_group_id", None)

            if existing_record:
                existing_record.plant_id = plant_id
                existing_record.equipment_id = equipment_id
                existing_record.equipment_group_id = group_id
                existing_record.created_by_user_id = current_user.id
                for field, value in row.items():
                    setattr(existing_record, field, value)
                updated += 1
            else:
                db_record = MaintenanceRecord(
                    plant_id=plant_id,
                    equipment_id=equipment_id,
                    equipment_group_id=group_id,
                    created_by_user_id=current_user.id,
                    **row,
                )
                db.add(db_record)
                created += 1

            saved += 1
        except Exception as exc:
            errors.append({"row": i + 1, "error": str(exc)})

    db.commit()
    log_action(db, current_user.id, "import", "maintenance_record", None, f"Imported {saved} records ({created} created, {updated} updated)")
    return {
        "saved": saved,
        "created": created,
        "updated": updated,
        "errors": errors,
        "message": f"Successfully imported {saved} records ({created} created, {updated} updated)",
    }
